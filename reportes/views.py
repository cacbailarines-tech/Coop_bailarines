from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Count, Q, Max
from socios.models import Socio, Libreta, AporteMensual, Periodo
from creditos.models import Credito
from multas.models import Multa
from core.utils import require_roles
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def reportes_index(request):
    return render(request, 'reportes/index.html')


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def reporte_cartera(request):
    periodo_id = request.GET.get('periodo')
    creditos = Credito.objects.select_related('socio','libreta')
    if periodo_id:
        creditos = creditos.filter(periodo_id=periodo_id)
    creditos = creditos.exclude(estado__in=['pendiente','cancelado'])
    total_activo = creditos.filter(estado__in=['desembolsado','mora_leve','mora_media','mora_grave']).aggregate(t=Sum('saldo_pendiente'))['t'] or 0
    total_mora = creditos.filter(estado__in=['mora_leve','mora_media','mora_grave']).aggregate(t=Sum('saldo_pendiente'))['t'] or 0
    total_recuperado = creditos.filter(estado='pagado').aggregate(t=Sum('monto_solicitado'))['t'] or 0
    periodos = Periodo.objects.all()
    return render(request, 'reportes/cartera.html', {
        'creditos': creditos, 'total_activo': total_activo,
        'total_mora': total_mora, 'total_recuperado': total_recuperado,
        'periodos': periodos, 'periodo_id': periodo_id,
    })


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def reporte_morosidad(request):
    creditos = Credito.objects.select_related('socio','libreta').filter(estado__in=['mora_leve','mora_media','mora_grave'])
    total = creditos.aggregate(t=Sum('saldo_pendiente'))['t'] or 0
    return render(request, 'reportes/morosidad.html', {'creditos': creditos, 'total': total})


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def reporte_ahorros(request):
    periodo_id = request.GET.get('periodo')
    libretas = Libreta.objects.select_related('socio','periodo').annotate(
        aportes_verificados_count=Count('aportes', filter=Q(aportes__estado='verificado')),
        ultimo_aporte_verificado=Max('aportes__fecha_verificacion', filter=Q(aportes__estado='verificado')),
    ).order_by('numero')
    if periodo_id:
        libretas = libretas.filter(periodo_id=periodo_id)
    total = libretas.aggregate(t=Sum('saldo_ahorro'))['t'] or 0
    periodos = Periodo.objects.all()
    return render(request, 'reportes/ahorros.html', {'libretas': libretas, 'total': total, 'periodos': periodos, 'periodo_id': periodo_id})


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def reporte_multas(request):
    multas = Multa.objects.select_related('socio','libreta','periodo').filter(estado='pendiente')
    total = multas.aggregate(t=Sum('monto'))['t'] or 0
    return render(request, 'reportes/multas.html', {'multas': multas, 'total': total})


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def reporte_aportes(request):
    periodo_id = request.GET.get('periodo')
    mes = request.GET.get('mes','')
    aportes = AporteMensual.objects.select_related('libreta__socio','libreta__periodo')
    if periodo_id:
        aportes = aportes.filter(libreta__periodo_id=periodo_id)
    if mes:
        aportes = aportes.filter(mes=int(mes))
    periodos = Periodo.objects.all()
    pendientes = aportes.filter(estado__in=['pendiente','atrasado']).count()
    verificados = aportes.filter(estado='verificado').count()
    return render(request, 'reportes/aportes.html', {
        'aportes': aportes, 'pendientes': pendientes, 'verificados': verificados,
        'periodos': periodos, 'periodo_id': periodo_id, 'mes': mes,
    })


def make_excel(title, headers, rows, fill_color="1B4F72"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    hfill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = Alignment(horizontal='center')
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    for col in ws.columns:
        mx = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(mx+4, 45)
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def exportar_socios(request):
    socios = Socio.objects.all()
    headers = ['Cédula','Nombres','Apellidos','Teléfono','WhatsApp','Email','Ciudad','Estado','Registro']
    rows = [(s.cedula,s.nombres,s.apellidos,s.telefono,s.whatsapp,s.email,s.ciudad,s.get_estado_display(),str(s.fecha_registro)) for s in socios]
    buf = make_excel('Socios', headers, rows)
    r = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    r['Content-Disposition'] = 'attachment; filename="socios_bailarines.xlsx"'
    return r


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def exportar_creditos(request):
    creditos = Credito.objects.select_related('socio','libreta').all()
    headers = ['N° Crédito','Socio','Cédula','Libreta','Tipo','Banco','Monto','Interés Total','Comisión','Transfiere','Cuota','Saldo','Estado','Fecha']
    rows = [(c.numero,c.socio.nombre_completo,c.socio.cedula,f'#{c.libreta.numero}',c.get_tipo_display(),c.get_banco_display(),float(c.monto_solicitado),float(c.interes_total),float(c.comision_bancaria),float(c.monto_transferir),float(c.cuota_mensual),float(c.saldo_pendiente),c.get_estado_display(),str(c.fecha_solicitud.date())) for c in creditos]
    buf = make_excel('Créditos', headers, rows, "922B21")
    r = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    r['Content-Disposition'] = 'attachment; filename="creditos_bailarines.xlsx"'
    return r


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def exportar_multas(request):
    multas = Multa.objects.select_related('socio','libreta').all()
    headers = ['Socio','Cédula','Libreta','Origen','Descripción','Monto','Estado','Fecha']
    rows = [(m.socio.nombre_completo,m.socio.cedula,f'#{m.libreta.numero}' if m.libreta else '-',m.get_origen_display(),m.descripcion,float(m.monto),m.get_estado_display(),str(m.fecha_generacion)) for m in multas]
    buf = make_excel('Multas', headers, rows, "6E2C00")
    r = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    r['Content-Disposition'] = 'attachment; filename="multas_bailarines.xlsx"'
    return r


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def reporte_movimientos(request):
    from core.models import Movimiento
    tipo = request.GET.get('tipo','')
    categoria = request.GET.get('categoria','')
    desde = request.GET.get('desde','')
    hasta = request.GET.get('hasta','')

    movs = Movimiento.objects.select_related('registrado_por').all()
    if tipo:
        movs = movs.filter(tipo=tipo)
    if categoria:
        movs = movs.filter(categoria=categoria)
    if desde:
        movs = movs.filter(fecha__gte=desde)
    if hasta:
        movs = movs.filter(fecha__lte=hasta)

    from django.db.models import Sum
    total_ingresos = movs.filter(tipo='ingreso').aggregate(t=Sum('monto'))['t'] or 0
    total_egresos = movs.filter(tipo='egreso').aggregate(t=Sum('monto'))['t'] or 0
    saldo = total_ingresos - total_egresos

    from core.models import Movimiento as M
    return render(request, 'reportes/movimientos.html', {
        'movs': movs, 'tipo': tipo, 'categoria': categoria,
        'desde': desde, 'hasta': hasta,
        'total_ingresos': total_ingresos, 'total_egresos': total_egresos, 'saldo': saldo,
        'categorias': M.CATEGORIA_CHOICES,
    })


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def exportar_movimientos(request):
    from core.models import Movimiento
    movs = Movimiento.objects.all()
    headers = ['Fecha','Tipo','Categoría','Descripción','Socio','Libreta','Crédito','Comprobante','Monto','Origen']
    rows = [(str(m.fecha), m.get_tipo_display(), m.get_categoria_display(), m.descripcion,
             m.socio_ref, m.libreta_ref, m.credito_ref, m.comprobante,
             float(m.monto), m.get_origen_display()) for m in movs]
    buf = make_excel('Movimientos', headers, rows, "1A5276")
    r = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    r['Content-Disposition'] = 'attachment; filename="movimientos_bailarines.xlsx"'
    return r


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def reporte_general(request):
    """Reporte completo de todos los movimientos del sistema"""
    from django.utils import timezone
    from socios.models import Socio, Libreta, AporteMensual, Periodo
    from creditos.models import Credito, PagoCredito
    from multas.models import Multa
    from core.models import Movimiento
    from cuentas.models import RifaMensual

    periodo_id = request.GET.get('periodo', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')

    # Base filters
    periodo_activo = Periodo.objects.filter(activo=True).first()
    periodo_sel = None
    if periodo_id:
        try:
            periodo_sel = Periodo.objects.get(pk=periodo_id)
        except:
            pass

    # === SOCIOS ===
    socios_total = Socio.objects.filter(estado='activo').count()
    socios_nuevos = Socio.objects.filter(estado='activo').count()

    # === LIBRETAS ===
    libretas_qs = Libreta.objects.select_related('socio', 'periodo')
    if periodo_sel:
        libretas_qs = libretas_qs.filter(periodo=periodo_sel)
    total_saldo_ahorro = libretas_qs.aggregate(t=Sum('saldo_ahorro'))['t'] or 0
    libretas_activas = libretas_qs.filter(estado='activa').count()

    # === APORTES ===
    aportes_qs = AporteMensual.objects.select_related('libreta__socio')
    if periodo_sel:
        aportes_qs = aportes_qs.filter(libreta__periodo=periodo_sel)
    if desde:
        aportes_qs = aportes_qs.filter(anio__gte=int(desde[:4]))
    aportes_verificados = aportes_qs.filter(estado='verificado')
    aportes_pendientes = aportes_qs.filter(estado__in=['pendiente','atrasado'])
    total_aportes = aportes_verificados.aggregate(t=Sum('monto_total'))['t'] or 0

    # === CRÉDITOS ===
    creditos_qs = Credito.objects.select_related('socio', 'libreta')
    if periodo_sel:
        creditos_qs = creditos_qs.filter(periodo=periodo_sel)
    total_desembolsado = creditos_qs.filter(estado__in=['desembolsado','mora_leve','mora_media','mora_grave','pagado']).aggregate(t=Sum('monto_transferir'))['t'] or 0
    cartera_activa = creditos_qs.filter(estado__in=['desembolsado','mora_leve','mora_media','mora_grave']).aggregate(t=Sum('saldo_pendiente'))['t'] or 0
    cartera_mora = creditos_qs.filter(estado__in=['mora_leve','mora_media','mora_grave']).aggregate(t=Sum('saldo_pendiente'))['t'] or 0
    creditos_pagados = creditos_qs.filter(estado='pagado').count()
    total_intereses = creditos_qs.aggregate(t=Sum('interes_total'))['t'] or 0

    # === PAGOS CRÉDITO ===
    pagos_qs = PagoCredito.objects.select_related('credito__socio').filter(estado='verificado')
    if periodo_sel:
        pagos_qs = pagos_qs.filter(credito__periodo=periodo_sel)
    total_cobrado = pagos_qs.aggregate(t=Sum('monto_pagado'))['t'] or 0

    # === MULTAS ===
    multas_qs = Multa.objects.select_related('socio', 'libreta')
    if periodo_sel:
        multas_qs = multas_qs.filter(periodo=periodo_sel)
    multas_pendientes_monto = multas_qs.filter(estado='pendiente').aggregate(t=Sum('monto'))['t'] or 0
    multas_cobradas = multas_qs.filter(estado='pagada').aggregate(t=Sum('monto'))['t'] or 0

    # === MOVIMIENTOS ===
    movs_qs = Movimiento.objects.all()
    if desde:
        movs_qs = movs_qs.filter(fecha__gte=desde)
    if hasta:
        movs_qs = movs_qs.filter(fecha__lte=hasta)
    total_ingresos = movs_qs.filter(tipo='ingreso').aggregate(t=Sum('monto'))['t'] or 0
    total_egresos = movs_qs.filter(tipo='egreso').aggregate(t=Sum('monto'))['t'] or 0
    saldo_neto = total_ingresos - total_egresos

    # Ingresos por categoria
    from django.db.models import Sum as DSum
    ingresos_por_cat = {}
    for mov in movs_qs.filter(tipo='ingreso').values('categoria').annotate(total=DSum('monto')):
        ingresos_por_cat[mov['categoria']] = mov['total']
    egresos_por_cat = {}
    for mov in movs_qs.filter(tipo='egreso').values('categoria').annotate(total=DSum('monto')):
        egresos_por_cat[mov['categoria']] = mov['total']

    # === RIFAS ===
    rifas_ganadas = RifaMensual.objects.filter(estado='ganada').count()
    rifas_total = RifaMensual.objects.count()

    # Recent activity lists
    ultimos_creditos = Credito.objects.select_related('socio','libreta').order_by('-fecha_solicitud')[:10]
    ultimas_multas = Multa.objects.select_related('socio').filter(estado='pendiente').order_by('-fecha_generacion')[:10]
    ultimos_aportes_pend = AporteMensual.objects.select_related('libreta__socio').filter(estado__in=['pendiente','atrasado']).order_by('anio','mes')[:10]

    periodos = Periodo.objects.all().order_by('-anio')

    return render(request, 'reportes/general.html', {
        'periodo_sel': periodo_sel, 'periodos': periodos,
        'periodo_id': periodo_id, 'desde': desde, 'hasta': hasta,
        # Socios
        'socios_total': socios_total,
        'libretas_activas': libretas_activas,
        'total_saldo_ahorro': total_saldo_ahorro,
        # Aportes
        'aportes_verificados': aportes_verificados.count(),
        'aportes_pendientes': aportes_pendientes.count(),
        'total_aportes': total_aportes,
        # Créditos
        'total_desembolsado': total_desembolsado,
        'cartera_activa': cartera_activa,
        'cartera_mora': cartera_mora,
        'creditos_pagados': creditos_pagados,
        'total_intereses': total_intereses,
        'total_cobrado': total_cobrado,
        # Multas
        'multas_pendientes_monto': multas_pendientes_monto,
        'multas_cobradas': multas_cobradas,
        # Movimientos
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'saldo_neto': saldo_neto,
        'ingresos_por_cat': ingresos_por_cat,
        'egresos_por_cat': egresos_por_cat,
        # Rifas
        'rifas_ganadas': rifas_ganadas,
        'rifas_total': rifas_total,
        # Lists
        'ultimos_creditos': ultimos_creditos,
        'ultimas_multas': ultimas_multas,
        'ultimos_aportes_pend': ultimos_aportes_pend,
    })


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def exportar_general_excel(request):
    """Exportar reporte general completo a Excel con múltiples hojas"""
    from socios.models import Socio, Libreta, AporteMensual
    from creditos.models import Credito, PagoCredito
    from multas.models import Multa
    from core.models import Movimiento
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()

    azul = 'FF0B3D91'
    verde = 'FF00897B'
    rojo = 'FFC62828'
    amarillo = 'FFF57F17'
    rosa = 'FFE91E63'

    def make_sheet(ws, title, headers, rows, color):
        ws.title = title
        hfill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        hfont = Font(color='FFFFFFFF', bold=True)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal='center')
        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
        for col in ws.columns:
            mx = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 50)

    # Hoja 1: Socios
    ws1 = wb.active
    socios = Socio.objects.all()
    make_sheet(ws1, 'Socios', ['Cédula','Nombres','Apellidos','Teléfono','WhatsApp','Email','Ciudad','Estado','Registro'],
        [(s.cedula,s.nombres,s.apellidos,s.telefono,s.whatsapp,s.email,s.ciudad,s.get_estado_display(),str(s.fecha_registro)) for s in socios], azul)

    # Hoja 2: Libretas
    ws2 = wb.create_sheet()
    libretas = Libreta.objects.select_related('socio','periodo').all()
    make_sheet(ws2, 'Libretas', ['#','Socio','Cédula','Periodo','Estado','Saldo Ahorro','Inscripción Pagada','Aportes Verif.'],
        [(lib.numero,lib.socio.nombre_completo,lib.socio.cedula,lib.periodo.anio,lib.get_estado_display(),
          float(lib.saldo_ahorro),'Sí' if lib.inscripcion_pagada else 'No',
          lib.aportes.filter(estado='verificado').count()) for lib in libretas], verde)

    # Hoja 3: Aportes
    ws3 = wb.create_sheet()
    aportes = AporteMensual.objects.select_related('libreta__socio').all().order_by('anio','mes')
    make_sheet(ws3, 'Aportes', ['Libreta #','Socio','Mes','Año','Ahorro','Lotería','Cumpleaños','Total','Estado','Comprobante'],
        [(a.libreta.numero,a.libreta.socio.nombre_completo,a.get_mes_display(),a.anio,
          float(a.monto_ahorro),float(a.monto_loteria),float(a.monto_cumpleanos),float(a.monto_total),
          a.get_estado_display(),a.comprobante_referencia) for a in aportes], verde)

    # Hoja 4: Créditos
    ws4 = wb.create_sheet()
    creditos = Credito.objects.select_related('socio','libreta').all()
    make_sheet(ws4, 'Créditos', ['N° Crédito','Socio','Cédula','Libreta','Tipo','Banco','Monto','Interés','Comisión','Transfiere','Cuota','Saldo','Estado','Fecha Solicitud','Fecha Desembolso','Vencimiento'],
        [(c.numero,c.socio.nombre_completo,c.socio.cedula,f'#{c.libreta.numero}',c.get_tipo_display(),c.get_banco_display(),
          float(c.monto_solicitado),float(c.interes_total),float(c.comision_bancaria),float(c.monto_transferir),
          float(c.cuota_mensual),float(c.saldo_pendiente),c.get_estado_display(),
          str(c.fecha_solicitud.date()),str(c.fecha_desembolso) if c.fecha_desembolso else '',
          str(c.fecha_pago_limite) if c.fecha_pago_limite else '') for c in creditos], amarillo)

    # Hoja 5: Pagos crédito
    ws5 = wb.create_sheet()
    pagos = PagoCredito.objects.select_related('credito__socio').all()
    make_sheet(ws5, 'Pagos Crédito', ['Crédito','Socio','# Pago','Monto','Comprobante','Estado','Fecha'],
        [(p.credito.numero,p.credito.socio.nombre_completo,p.numero_pago,float(p.monto_pagado),
          p.comprobante_referencia,p.get_estado_display(),str(p.fecha_reporte.date())) for p in pagos], amarillo)

    # Hoja 6: Multas
    ws6 = wb.create_sheet()
    multas = Multa.objects.select_related('socio','libreta').all()
    make_sheet(ws6, 'Multas', ['Socio','Cédula','Libreta','Origen','Descripción','Monto','Estado','Fecha'],
        [(m.socio.nombre_completo,m.socio.cedula,f'#{m.libreta.numero}' if m.libreta else '-',
          m.get_origen_display(),m.descripcion,float(m.monto),m.get_estado_display(),str(m.fecha_generacion)) for m in multas], rojo)

    # Hoja 7: Movimientos (Ingresos/Egresos)
    ws7 = wb.create_sheet()
    movs = Movimiento.objects.all()
    make_sheet(ws7, 'Ingresos y Egresos', ['Fecha','Tipo','Categoría','Descripción','Socio','Libreta','Crédito','Comprobante','Monto','Origen'],
        [(str(m.fecha),m.get_tipo_display(),m.get_categoria_display(),m.descripcion,
          m.socio_ref,m.libreta_ref,m.credito_ref,m.comprobante,float(m.monto),m.get_origen_display()) for m in movs], azul)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_general_bailarines.xlsx"'
    return response
