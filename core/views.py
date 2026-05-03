import os
from datetime import date as ddate
from decimal import Decimal

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.views import LoginView as BaseLoginView
from django.core.management import call_command
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse, FileResponse
from django.shortcuts import get_object_or_404, redirect, render
import io
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from creditos.models import Credito, PagoCredito
from multas.models import Multa
from socios.models import AporteMensual, Libreta, Periodo, Socio

from .email_notifications import (
    notify_aporte_rechazado,
    notify_aporte_verificado,
    notify_cumpleanos_transferido,
    notify_multa_rechazada,
    notify_multa_verificada,
    notify_pago_credito_rechazado,
    notify_pago_credito_verificado,
)
from .models import AuditoriaSistema, CajaDiaria, Movimiento, PerfilUsuario
from .utils import has_role, registrar_auditoria, require_roles


def home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'core/landing.html')


def about(request):
    return render(request, 'core/about.html')


def redirect_root(request):
    return redirect('dashboard') if request.user.is_authenticated else redirect('login')


@login_required
@require_roles('admin')
def descargar_respaldo_bd(request):
    if not request.user.is_superuser:
        messages.error(request, 'Solo los superusuarios pueden descargar respaldos.')
        return redirect('dashboard')
    
    out = io.StringIO()
    call_command('dumpdata', exclude=['contenttypes', 'auth.permission', 'sessions', 'admin.logentry'], format='json', indent=2, stdout=out)
    
    response = HttpResponse(out.getvalue(), content_type='application/json')
    fecha_str = timezone.localdate().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="respaldo_cooperativa_{fecha_str}.json"'
    
    registrar_auditoria(request.user, 'seguridad', 'descargar_respaldo', 'Se descargo un respaldo completo de la base de datos en formato JSON.')
    return response


def pwa_manifest(request):
    return JsonResponse({
        "name": "Cooperativa Bailarines",
        "short_name": "Bailarines",
        "description": "Sistema web administrativo y portal del socio de la Cooperativa Bailarines.",
        "start_url": "/portal/",
        "scope": "/",
        "display": "standalone",
        "background_color": "#F0F4FA",
        "theme_color": "#0B3D91",
        "orientation": "portrait",
        "icons": [
            {
                "src": "/static/img/logo.png",
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "any"
            },
            {
                "src": "/static/img/logo.png",
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "any"
            }
        ],
        "shortcuts": [
            {
                "name": "Portal del Socio",
                "short_name": "Portal",
                "url": "/portal/",
                "icons": [{"src": "/static/img/logo.png", "sizes": "192x192", "type": "image/png"}]
            },
            {
                "name": "Panel Principal",
                "short_name": "Dashboard",
                "url": "/dashboard/",
                "icons": [{"src": "/static/img/logo.png", "sizes": "192x192", "type": "image/png"}]
            }
        ],
        "share_target": {
            "action": "/portal/compartir/",
            "method": "POST",
            "enctype": "multipart/form-data",
            "params": {
                "title": "title",
                "text": "text",
                "files": [
                    {
                        "name": "comprobante_archivo",
                        "accept": ["image/*", "application/pdf"]
                    }
                ]
            }
        }
    })


def service_worker(request):
    import time
    response = render(request, 'pwa/service-worker.js', {
        'static_version': f'v-{int(time.time())}',
    }, content_type='application/javascript')
    response['Service-Worker-Allowed'] = '/'
    return response


def offline_page(request):
    return render(request, 'pwa/offline.html')


def push_config(request):
    return JsonResponse({
        'enabled': bool(settings.WEBPUSH_ENABLED and settings.WEBPUSH_VAPID_PUBLIC_KEY),
        'publicKey': settings.WEBPUSH_VAPID_PUBLIC_KEY,
    })


class LoginView(BaseLoginView):
    template_name = 'core/login.html'
    redirect_authenticated_user = True

    def form_valid(self, form):
        response = super().form_valid(form)
        keep_signed_in = self.request.POST.get('keep_signed_in')
        self.request.session.set_expiry(settings.SESSION_COOKIE_AGE if keep_signed_in else 0)
        return response


def logout_view(request):
    logout(request)
    return redirect('login')


def bootstrap_admin(request):
    token = os.getenv('BOOTSTRAP_ADMIN_TOKEN', '').strip()
    has_superuser = User.objects.filter(is_superuser=True).exists()
    context = {
        'token_configured': bool(token),
        'has_superuser': has_superuser,
    }

    if request.method == 'POST':
        submitted_token = request.POST.get('token', '').strip()
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')

        context.update({
            'username': username,
            'email': email,
        })

        if not token:
            messages.error(request, 'La ruta de creación inicial está deshabilitada.')
        elif submitted_token != token:
            messages.error(request, 'Token inválido.')
        elif has_superuser:
            messages.warning(request, 'Ya existe un superusuario. Desactive esta ruta quitando BOOTSTRAP_ADMIN_TOKEN.')
        elif not username or not password:
            messages.error(request, 'Debe completar usuario y contraseña.')
        elif password != password_confirm:
            messages.error(request, 'Las contraseñas no coinciden.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Ese nombre de usuario ya existe.')
        else:
            User.objects.create_superuser(
                username=username,
                email=email,
                password=password,
            )
            messages.success(request, 'Superusuario creado correctamente. Ya puede iniciar sesión.')
            return redirect('login')

    return render(request, 'core/bootstrap_admin.html', context)


def _badge_estado_credito(estado):
    if estado == 'pagado':
        return 'badge-success'
    if 'mora' in estado:
        return 'badge-danger'
    if estado == 'desembolsado':
        return 'badge-primary'
    if estado == 'aprobado':
        return 'badge-warning'
    return 'badge-secondary'


def _build_excel_sheet(ws, headers, rows, color='0B3D91'):
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    font = Font(color='FFFFFF', bold=True)
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col in ws.columns:
        width = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 4, 45)


@login_required
def dashboard(request):
    periodo_activo = Periodo.objects.filter(activo=True).first()
    hoy = timezone.localdate()
    total_socios = Socio.objects.filter(estado='activo').count()
    total_libretas = Libreta.objects.filter(estado='activa').count()
    if periodo_activo:
        total_libretas = Libreta.objects.filter(periodo=periodo_activo, estado='activa').count()
        total_ahorros = Libreta.objects.filter(periodo=periodo_activo).aggregate(t=Sum('saldo_ahorro'))['t'] or 0
    else:
        total_ahorros = Libreta.objects.aggregate(t=Sum('saldo_ahorro'))['t'] or 0

    creditos_activos = Credito.objects.filter(
        estado__in=['desembolsado', 'mora_leve', 'mora_media', 'mora_grave']
    )
    total_cartera = creditos_activos.aggregate(t=Sum('saldo_pendiente'))['t'] or 0
    creditos_mora = Credito.objects.filter(
        estado__in=['mora_leve', 'mora_media', 'mora_grave']
    ).count()
    multas_pendientes = Multa.objects.filter(estado='pendiente').aggregate(t=Sum('monto'))['t'] or 0
    aportes_por_verificar = AporteMensual.objects.filter(estado='reportado').count()
    pagos_por_verificar = PagoCredito.objects.filter(estado='reportado').count()
    multas_por_verificar = Multa.objects.filter(
        estado='pendiente',
        observaciones__startswith='Pago reportado'
    ).count()
    ultimos_creditos = Credito.objects.select_related('socio', 'libreta').order_by('-fecha_solicitud')[:6]
    ultimas_multas = Multa.objects.select_related('socio').filter(estado='pendiente').order_by('-fecha_generacion')[:6]
    creditos_por_vencer_qs = Credito.objects.filter(
        estado__in=['desembolsado', 'mora_leve', 'mora_media', 'mora_grave'],
        saldo_pendiente__gt=0,
        fecha_pago_limite__isnull=False,
        fecha_pago_limite__gte=hoy,
        fecha_pago_limite__lte=hoy + timezone.timedelta(days=15),
    ).select_related('socio', 'libreta').order_by('fecha_pago_limite')[:8]
    creditos_por_vencer = []
    for credito in creditos_por_vencer_qs:
        dias_restantes = (credito.fecha_pago_limite - hoy).days
        if dias_restantes == 0:
            nivel = 'hoy'
            badge = 'badge-danger'
            texto = 'Vence hoy'
        elif dias_restantes == 1:
            nivel = 'manana'
            badge = 'badge-warning'
            texto = 'Vence ma\xf1ana'
        elif dias_restantes <= 3:
            nivel = 'tres_dias'
            badge = 'badge-warning'
            texto = f'Vence en {dias_restantes} d\xedas'
        else:
            nivel = 'semana'
            badge = 'badge-primary'
            texto = f'Vence en {dias_restantes} d\xedas'
        creditos_por_vencer.append({
            'credito': credito,
            'dias_restantes': dias_restantes,
            'texto_alerta': texto,
            'badge': badge,
            'nivel': nivel,
        })
    resumen_vencimientos = {
        'hoy': sum(1 for item in creditos_por_vencer if item['nivel'] == 'hoy'),
        'manana': sum(1 for item in creditos_por_vencer if item['nivel'] == 'manana'),
        'tres_dias': sum(1 for item in creditos_por_vencer if item['nivel'] == 'tres_dias'),
        'semana': sum(1 for item in creditos_por_vencer if item['nivel'] == 'semana'),
    }

    return render(request, 'core/dashboard.html', {
        'periodo_activo': periodo_activo,
        'hoy': hoy,
        'total_socios': total_socios,
        'total_libretas': total_libretas,
        'total_ahorros': total_ahorros,
        'total_cartera': total_cartera,
        'creditos_mora': creditos_mora,
        'multas_pendientes': multas_pendientes,
        'aportes_por_verificar': aportes_por_verificar,
        'pagos_por_verificar': pagos_por_verificar,
        'multas_por_verificar': multas_por_verificar,
        'ultimos_creditos': ultimos_creditos,
        'ultimas_multas': ultimas_multas,
        'creditos_por_vencer': creditos_por_vencer,
        'resumen_vencimientos': resumen_vencimientos,
        'es_tesoreria': has_role(request.user, 'admin', 'tesorero', 'gerente'),
    })


@login_required
@require_roles('admin', 'gerente')
def periodos_list(request):
    return render(request, 'core/periodos_list.html', {'periodos': Periodo.objects.all()})


@login_required
@require_roles('admin', 'gerente')
def periodo_crear(request):
    if request.method == 'POST':
        anio = int(request.POST.get('anio'))
        if Periodo.objects.filter(anio=anio).exists():
            messages.error(request, f'Ya existe el periodo {anio}.')
        else:
            if request.POST.get('activo') == 'on':
                Periodo.objects.update(activo=False)
            periodo = Periodo.objects.create(
                anio=anio,
                fecha_inicio=request.POST.get('fecha_inicio'),
                fecha_cierre=request.POST.get('fecha_cierre'),
                activo=request.POST.get('activo') == 'on',
                descripcion=request.POST.get('descripcion', ''),
            )
            registrar_auditoria(request.user, 'seguridad', 'crear_periodo', f'Se cre? el periodo {anio}.', 'Periodo', periodo.pk)
            messages.success(request, f'Periodo {anio} creado.')
            return redirect('periodos_list')
    return render(request, 'core/periodo_form.html')


@login_required
@require_roles('admin')
def usuarios_list(request):
    usuarios = User.objects.select_related('perfil').all()
    return render(request, 'core/usuarios_list.html', {'usuarios': usuarios})


@login_required
@require_roles('admin')
def usuario_crear(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        if User.objects.filter(username=username).exists():
            messages.error(request, 'El usuario ya existe.')
        else:
            user = User.objects.create_user(
                username=username,
                password=request.POST.get('password'),
                first_name=request.POST.get('first_name', ''),
                last_name=request.POST.get('last_name', ''),
                email=request.POST.get('email', '')
            )
            PerfilUsuario.objects.create(usuario=user, rol=request.POST.get('rol', 'cajero'), telefono=request.POST.get('telefono', ''))
            registrar_auditoria(request.user, 'seguridad', 'crear_usuario', f'Se cre? el usuario {username}.', 'User', user.pk)
            messages.success(request, f'Usuario {username} creado.')
            return redirect('usuarios_list')
    return render(request, 'core/usuario_form.html', {'accion': 'Crear'})


@login_required
@require_roles('admin')
def usuario_editar(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        pwd = request.POST.get('password')
        if pwd:
            user.set_password(pwd)
        user.save()
        perfil, _ = PerfilUsuario.objects.get_or_create(usuario=user)
        perfil.rol = request.POST.get('rol', perfil.rol)
        perfil.telefono = request.POST.get('telefono', perfil.telefono)
        perfil.save()
        registrar_auditoria(request.user, 'seguridad', 'editar_usuario', f'Se actualiz? el usuario {user.username}.', 'User', user.pk, {'rol': perfil.rol})
        messages.success(request, 'Usuario actualizado.')
        return redirect('usuarios_list')
    return render(request, 'core/usuario_form.html', {'accion': 'Editar', 'obj': user})


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def tesoreria_dashboard(request):
    hoy = timezone.localdate()
    inicio_mes = hoy.replace(day=1)

    caja_hoy = CajaDiaria.objects.filter(fecha=hoy).first()
    ingresos_hoy = Movimiento.objects.filter(fecha=hoy, tipo='ingreso').aggregate(t=Sum('monto'))['t'] or 0
    egresos_hoy = Movimiento.objects.filter(fecha=hoy, tipo='egreso').aggregate(t=Sum('monto'))['t'] or 0
    saldo_hoy = (caja_hoy.saldo_inicial if caja_hoy else 0) + ingresos_hoy - egresos_hoy

    movimientos_mes = Movimiento.objects.filter(fecha__gte=inicio_mes, fecha__lte=hoy)
    ingresos_mes = movimientos_mes.filter(tipo='ingreso').aggregate(t=Sum('monto'))['t'] or 0
    egresos_mes = movimientos_mes.filter(tipo='egreso').aggregate(t=Sum('monto'))['t'] or 0

    aportes_pendientes = AporteMensual.objects.filter(estado='reportado')
    pagos_pendientes = PagoCredito.objects.filter(estado='reportado')
    multas_reportadas = Multa.objects.filter(
        estado='pendiente',
        observaciones__startswith='Pago reportado'
    )

    alertas = {
        'pendientes_antiguos': (
            aportes_pendientes.filter(fecha_reporte__date__lt=hoy).count() +
            pagos_pendientes.filter(fecha_reporte__date__lt=hoy).count()
        ),
        'multas_antiguas': Multa.objects.filter(
            estado='pendiente', fecha_generacion__lt=hoy - timezone.timedelta(days=30)
        ).count(),
        'creditos_mora': Credito.objects.filter(
            estado__in=['mora_leve', 'mora_media', 'mora_grave']
        ).count(),
        'creditos_por_vencer': Credito.objects.filter(
            estado__in=['desembolsado', 'mora_leve', 'mora_media', 'mora_grave'],
            saldo_pendiente__gt=0,
            fecha_pago_limite__isnull=False,
            fecha_pago_limite__gte=hoy,
            fecha_pago_limite__lte=hoy + timezone.timedelta(days=15),
        ).count(),
        'caja_abierta': bool(caja_hoy and caja_hoy.esta_abierta),
    }
    creditos_por_vencer_qs = list(
        Credito.objects.filter(
            estado__in=['desembolsado', 'mora_leve', 'mora_media', 'mora_grave'],
            saldo_pendiente__gt=0,
            fecha_pago_limite__isnull=False,
            fecha_pago_limite__gte=hoy,
            fecha_pago_limite__lte=hoy + timezone.timedelta(days=7),
        )
        .select_related('socio', 'libreta')
        .order_by('fecha_pago_limite')[:8]
    )
    creditos_por_vencer = []
    for credito in creditos_por_vencer_qs:
        dias_restantes = (credito.fecha_pago_limite - hoy).days
        if dias_restantes == 0:
            texto = 'Vence hoy'
            badge = 'badge-danger'
        elif dias_restantes == 1:
            texto = 'Vence mañana'
            badge = 'badge-warning'
        else:
            texto = f'Vence en {dias_restantes} días'
            badge = 'badge-primary'
        creditos_por_vencer.append({
            'credito': credito,
            'texto': texto,
            'badge': badge,
        })

    movimientos_recientes = Movimiento.objects.select_related('registrado_por').all()[:12]
    auditorias = AuditoriaSistema.objects.select_related('usuario').all()[:15]

    return render(request, 'core/tesoreria_dashboard.html', {
        'hoy': hoy,
        'caja_hoy': caja_hoy,
        'ingresos_hoy': ingresos_hoy,
        'egresos_hoy': egresos_hoy,
        'saldo_hoy': saldo_hoy,
        'ingresos_mes': ingresos_mes,
        'egresos_mes': egresos_mes,
        'saldo_mes': ingresos_mes - egresos_mes,
        'total_verificaciones': aportes_pendientes.count() + pagos_pendientes.count() + multas_reportadas.count(),
        'alertas': alertas,
        'aportes_pendientes': aportes_pendientes.count(),
        'pagos_pendientes': pagos_pendientes.count(),
        'multas_reportadas': multas_reportadas.count(),
        'creditos_por_vencer': creditos_por_vencer,
        'movimientos_recientes': movimientos_recientes,
        'auditorias': auditorias,
    })


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def caja_diaria_accion(request):
    if request.method != 'POST':
        return redirect('tesoreria_dashboard')

    hoy = timezone.localdate()
    accion = request.POST.get('accion')
    caja = CajaDiaria.objects.filter(fecha=hoy).first()

    if accion == 'abrir':
        if caja and caja.esta_abierta:
            messages.info(request, 'La caja del d?a ya est? abierta.')
        else:
            saldo_inicial = request.POST.get('saldo_inicial') or 0
            caja, creada = CajaDiaria.objects.get_or_create(
                fecha=hoy,
                defaults={
                    'saldo_inicial': saldo_inicial,
                    'abierta_por': request.user,
                    'observaciones_apertura': request.POST.get('observaciones_apertura', ''),
                }
            )
            if not creada:
                caja.saldo_inicial = saldo_inicial
                caja.abierta_por = request.user
                caja.fecha_apertura = timezone.now()
                caja.fecha_cierre = None
                caja.cerrada_por = None
                caja.observaciones_apertura = request.POST.get('observaciones_apertura', '')
                caja.observaciones_cierre = ''
                caja.save()
            registrar_auditoria(request.user, 'caja', 'abrir_caja', f'Se abri? la caja del {hoy}.', 'CajaDiaria', caja.pk, {'saldo_inicial': str(caja.saldo_inicial)})
            messages.success(request, 'Caja abierta correctamente.')
    elif accion == 'cerrar':
        if not caja or not caja.esta_abierta:
            messages.error(request, 'No hay una caja abierta para hoy.')
        else:
            caja.fecha_cierre = timezone.now()
            caja.cerrada_por = request.user
            caja.observaciones_cierre = request.POST.get('observaciones_cierre', '')
            caja.save()
            registrar_auditoria(request.user, 'caja', 'cerrar_caja', f'Se cerr? la caja del {hoy}.', 'CajaDiaria', caja.pk)
            messages.success(request, 'Caja cerrada correctamente.')

    return redirect('tesoreria_dashboard')


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def movimientos_list(request):
    tipo = request.GET.get('tipo', '')
    categoria = request.GET.get('categoria', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    origen = request.GET.get('origen', '')
    registrado_por = request.GET.get('registrado_por', '')
    monto_min = request.GET.get('monto_min', '')
    monto_max = request.GET.get('monto_max', '')
    q = request.GET.get('q', '')

    movs = Movimiento.objects.select_related('registrado_por').all()
    if tipo:
        movs = movs.filter(tipo=tipo)
    if categoria:
        movs = movs.filter(categoria=categoria)
    if desde:
        movs = movs.filter(fecha__gte=desde)
    if hasta:
        movs = movs.filter(fecha__lte=hasta)
    if origen:
        movs = movs.filter(origen=origen)
    if registrado_por:
        movs = movs.filter(registrado_por_id=registrado_por)
    if monto_min:
        movs = movs.filter(monto__gte=monto_min)
    if monto_max:
        movs = movs.filter(monto__lte=monto_max)
    if q:
        movs = movs.filter(
            Q(descripcion__icontains=q) |
            Q(comprobante__icontains=q) |
            Q(socio_ref__icontains=q) |
            Q(libreta_ref__icontains=q) |
            Q(credito_ref__icontains=q)
        )

    total_ingresos = movs.filter(tipo='ingreso').aggregate(t=Sum('monto'))['t'] or 0
    total_egresos = movs.filter(tipo='egreso').aggregate(t=Sum('monto'))['t'] or 0
    saldo = total_ingresos - total_egresos
    movs = movs.order_by('-fecha_registro', '-fecha', '-pk')
    paginator = Paginator(movs, 20)
    page = paginator.get_page(request.GET.get('page'))
    usuarios = User.objects.filter(movimientos__isnull=False).distinct().order_by('first_name', 'username')

    return render(request, 'core/movimientos_list.html', {
        'page_obj': page,
        'tipo': tipo,
        'categoria': categoria,
        'desde': desde,
        'hasta': hasta,
        'origen': origen,
        'registrado_por': registrado_por,
        'monto_min': monto_min,
        'monto_max': monto_max,
        'q': q,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'saldo': saldo,
        'categorias': Movimiento.CATEGORIA_CHOICES,
        'usuarios': usuarios,
    })


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def movimiento_crear(request):
    if request.method == 'POST':
        fecha_tipo = request.POST.get('fecha_tipo', 'hoy')
        fecha = timezone.localdate() if fecha_tipo == 'hoy' else request.POST.get('fecha_manual', str(timezone.localdate()))
        movimiento = Movimiento.objects.create(
            tipo=request.POST.get('tipo'),
            categoria=request.POST.get('categoria'),
            descripcion=request.POST.get('descripcion', ''),
            monto=request.POST.get('monto'),
            fecha=fecha,
            origen='manual',
            comprobante=request.POST.get('comprobante', ''),
            comprobante_archivo=request.FILES.get('comprobante_archivo'),
            socio_ref=request.POST.get('socio_ref', ''),
            libreta_ref=request.POST.get('libreta_ref', ''),
            credito_ref=request.POST.get('credito_ref', ''),
            registrado_por=request.user,
            observaciones=request.POST.get('observaciones', ''),
        )
        registrar_auditoria(request.user, 'tesoreria', 'crear_movimiento', f'Se registr? un {movimiento.get_tipo_display().lower()} manual por ${movimiento.monto}.', 'Movimiento', movimiento.pk, {'categoria': movimiento.categoria})
        messages.success(request, 'Movimiento registrado exitosamente.')
        return redirect('movimientos_list')

    return render(request, 'core/movimiento_form.html', {'categorias': Movimiento.CATEGORIA_CHOICES, 'accion': 'Registrar', 'hoy': timezone.localdate()})


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def movimiento_editar(request, pk):
    mov = get_object_or_404(Movimiento, pk=pk)
    if request.method == 'POST':
        fecha_tipo = request.POST.get('fecha_tipo', 'manual')
        mov.fecha = timezone.localdate() if fecha_tipo == 'hoy' else request.POST.get('fecha_manual', str(mov.fecha))
        mov.tipo = request.POST.get('tipo', mov.tipo)
        mov.categoria = request.POST.get('categoria', mov.categoria)
        mov.descripcion = request.POST.get('descripcion', mov.descripcion)
        mov.monto = request.POST.get('monto', mov.monto)
        mov.comprobante = request.POST.get('comprobante', '')
        if request.FILES.get('comprobante_archivo'):
            mov.comprobante_archivo = request.FILES.get('comprobante_archivo')
        mov.socio_ref = request.POST.get('socio_ref', '')
        mov.libreta_ref = request.POST.get('libreta_ref', '')
        mov.credito_ref = request.POST.get('credito_ref', '')
        mov.observaciones = request.POST.get('observaciones', '')
        mov.save()
        registrar_auditoria(request.user, 'tesoreria', 'editar_movimiento', f'Se edit? el movimiento #{mov.pk}.', 'Movimiento', mov.pk)
        messages.success(request, 'Movimiento actualizado.')
        return redirect('movimientos_list')

    return render(request, 'core/movimiento_form.html', {'categorias': Movimiento.CATEGORIA_CHOICES, 'accion': 'Editar', 'mov': mov, 'hoy': timezone.localdate()})


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def exportar_verificaciones_excel(request):
    wb = Workbook()
    ws_aportes = wb.active
    ws_aportes.title = 'Aportes'
    aportes = AporteMensual.objects.filter(estado='reportado').select_related('libreta__socio')
    _build_excel_sheet(
        ws_aportes,
        ['Libreta', 'Socio', 'Mes', 'A?o', 'Monto', 'Comprobante', 'Fecha', 'Adjunto'],
        [
            (
                a.libreta.numero,
                a.libreta.socio.nombre_completo,
                a.get_mes_display(),
                a.anio,
                float(a.monto_total),
                a.comprobante_referencia,
                a.fecha_reporte.strftime('%d/%m/%Y %H:%M') if a.fecha_reporte else '',
                request.build_absolute_uri(a.comprobante_archivo.url) if a.comprobante_archivo else '',
            )
            for a in aportes
        ],
        '0B3D91'
    )

    ws_pagos = wb.create_sheet('Pagos Cr?dito')
    pagos = PagoCredito.objects.filter(estado='reportado').select_related('credito__socio', 'credito__libreta')
    _build_excel_sheet(
        ws_pagos,
        ['Cr?dito', 'Socio', 'Libreta', 'Monto', 'Comprobante', 'Fecha', 'Adjunto'],
        [
            (
                p.credito.numero,
                p.credito.socio.nombre_completo,
                p.credito.libreta.numero,
                float(p.monto_pagado),
                p.comprobante_referencia,
                p.fecha_reporte.strftime('%d/%m/%Y %H:%M'),
                request.build_absolute_uri(p.comprobante_archivo.url) if p.comprobante_archivo else '',
            )
            for p in pagos
        ],
        'F57F17'
    )

    ws_multas = wb.create_sheet('Multas')
    multas = Multa.objects.filter(estado='pendiente', observaciones__startswith='Pago reportado').select_related('socio', 'libreta')
    _build_excel_sheet(
        ws_multas,
        ['Socio', 'Libreta', 'Monto', 'Descripci?n', 'Observaciones', 'Adjunto'],
        [
            (
                m.socio.nombre_completo,
                m.libreta.numero if m.libreta else '-',
                float(m.monto),
                m.descripcion,
                m.observaciones,
                request.build_absolute_uri(m.comprobante_archivo.url) if m.comprobante_archivo else '',
            )
            for m in multas
        ],
        'C62828'
    )

    registrar_auditoria(request.user, 'reportes', 'exportar_verificaciones', 'Se export? el panel de verificaciones a Excel.')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="verificaciones_pendientes.xlsx"'
    wb.save(response)
    return response


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def verificar_todo(request):
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        pk = request.POST.get('pk')
        accion = request.POST.get('accion')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '').strip()

        if accion == 'rechazar' and not observacion:
            messages.error(request, 'Debe ingresar una observaci?n para rechazar el reporte.')
            return redirect('verificar_todo')

        if tipo == 'aporte':
            aporte = get_object_or_404(AporteMensual, pk=pk)
            if accion == 'verificar':
                aporte.estado = 'verificado'
                aporte.fecha_verificacion = timezone.now()
                aporte.comprobante_referencia = comprobante or aporte.comprobante_referencia
                aporte.observacion = ''
                aporte.save()
                notify_aporte_verificado(aporte)
                lib = aporte.libreta
                lib.saldo_ahorro += aporte.monto_ahorro
                lib.save()
                Movimiento.objects.create(
                    tipo='ingreso',
                    categoria='aporte_mensual',
                    descripcion=f'Aporte {aporte.get_mes_display()} {aporte.anio} - Libreta #{lib.numero} - {lib.socio.nombre_completo}',
                    monto=aporte.monto_total,
                    fecha=timezone.localdate(),
                    origen='automatico',
                    comprobante=comprobante,
                    comprobante_archivo=aporte.comprobante_archivo,
                    socio_ref=lib.socio.nombre_completo,
                    libreta_ref=f'#{lib.numero}',
                    registrado_por=request.user,
                )
                registrar_auditoria(request.user, 'verificaciones', 'verificar_aporte', f'Se verific? el aporte {aporte.get_mes_display()} {aporte.anio}.', 'AporteMensual', aporte.pk)
                messages.success(request, f'Aporte de {aporte.get_mes_display()} verificado.')
            elif accion == 'rechazar':
                aporte.estado = 'pendiente'
                aporte.observacion = f'Rechazado: {observacion}'
                aporte.save()
                notify_aporte_rechazado(aporte, observacion)
                registrar_auditoria(request.user, 'verificaciones', 'rechazar_aporte', f'Se rechaz? el aporte {aporte.get_mes_display()} {aporte.anio}.', 'AporteMensual', aporte.pk, {'motivo': observacion})
                messages.warning(request, 'Aporte rechazado.')

        elif tipo == 'pago_credito':
            pago = get_object_or_404(PagoCredito, pk=pk)
            if accion == 'verificar':
                pago.estado = 'verificado'
                pago.fecha_verificacion = timezone.now()
                pago.verificado_por = request.user
                pago.observaciones = ''
                pago.save()
                notify_pago_credito_verificado(pago)
                credito = pago.credito
                tasa = credito.tasa_mensual
                saldo_antes = pago.saldo_anterior
                interes = saldo_antes * tasa
                capital = max(pago.monto_pagado - interes, 0)
                Movimiento.objects.create(
                    tipo='ingreso',
                    categoria='interes_credito',
                    descripcion=f'Pago #{pago.numero_pago} - cr?dito {credito.numero} - {credito.socio.nombre_completo}',
                    monto=pago.monto_pagado,
                    fecha=timezone.localdate(),
                    origen='automatico',
                    comprobante=pago.comprobante_referencia,
                    comprobante_archivo=pago.comprobante_archivo,
                    socio_ref=credito.socio.nombre_completo,
                    credito_ref=credito.numero,
                    registrado_por=request.user,
                    observaciones=f'Capital: ${capital:.2f} | Inter?s estimado: ${interes:.2f}',
                )
                registrar_auditoria(request.user, 'verificaciones', 'verificar_pago_credito', f'Se verific? el pago #{pago.numero_pago} del cr?dito {credito.numero}.', 'PagoCredito', pago.pk)
                messages.success(request, f'Pago #{pago.numero_pago} verificado.')
            elif accion == 'rechazar':
                pago.estado = 'rechazado'
                pago.observaciones = observacion
                pago.credito.saldo_pendiente = pago.saldo_anterior
                if pago.credito.estado == 'pagado':
                    pago.credito.estado = 'desembolsado'
                pago.credito.save()
                pago.save()
                registrar_auditoria(request.user, 'verificaciones', 'rechazar_pago_credito', f'Se rechaz? el pago #{pago.numero_pago} del cr?dito {pago.credito.numero}.', 'PagoCredito', pago.pk, {'motivo': observacion})
                notify_pago_credito_rechazado(pago, observacion)
                messages.warning(request, 'Pago rechazado. Saldo restaurado.')

        elif tipo == 'multa':
            multa = get_object_or_404(Multa, pk=pk)
            if accion == 'verificar':
                multa.estado = 'pagada'
                multa.fecha_pago = timezone.localdate()
                multa.comprobante_pago = comprobante
                multa.save()
                Movimiento.objects.create(
                    tipo='ingreso',
                    categoria='multa',
                    descripcion=f'Multa pagada - {multa.socio.nombre_completo}: {multa.descripcion}',
                    monto=multa.monto,
                    fecha=timezone.localdate(),
                    origen='automatico',
                    comprobante=comprobante,
                    comprobante_archivo=multa.comprobante_archivo,
                    socio_ref=multa.socio.nombre_completo,
                    libreta_ref=f'#{multa.libreta.numero}' if multa.libreta else '',
                    registrado_por=request.user,
                )
                registrar_auditoria(request.user, 'verificaciones', 'verificar_multa', f'Se verific? el pago de multa #{multa.pk}.', 'Multa', multa.pk)
                notify_multa_verificada(multa)
                messages.success(request, f'Multa de ${multa.monto} marcada como pagada.')
            elif accion == 'rechazar':
                multa.observaciones = f'Reporte rechazado: {observacion}'
                multa.comprobante_pago = ''
                multa.comprobante_archivo = None
                multa.save()
                registrar_auditoria(request.user, 'verificaciones', 'rechazar_multa', f'Se rechaz? el pago reportado de la multa #{multa.pk}.', 'Multa', multa.pk, {'motivo': observacion})
                notify_multa_rechazada(multa, observacion)
                messages.warning(request, 'Reporte de multa rechazado.')

        return redirect('verificar_todo')

    aportes_pendientes = AporteMensual.objects.filter(estado='reportado').select_related('libreta__socio', 'libreta__periodo').order_by('-fecha_reporte')
    pagos_credito = PagoCredito.objects.filter(estado='reportado').select_related('credito__socio', 'credito__libreta').order_by('-fecha_reporte')
    multas_reportadas = Multa.objects.filter(estado='pendiente', observaciones__startswith='Pago reportado').select_related('socio', 'libreta').order_by('-fecha_generacion')
    total = aportes_pendientes.count() + pagos_credito.count() + multas_reportadas.count()

    return render(request, 'core/verificar_todo.html', {
        'aportes_pendientes': aportes_pendientes,
        'pagos_credito': pagos_credito,
        'multas_reportadas': multas_reportadas,
        'total': total,
    })


@login_required
def solicitudes_pendientes(request):
    solicitudes = Credito.objects.filter(estado='pendiente').select_related('socio', 'libreta', 'periodo').order_by('fecha_solicitud')
    return render(request, 'core/solicitudes_pendientes.html', {'solicitudes': solicitudes})


@login_required
def cumpleanos_calendario(request):
    mes_sel = int(request.GET.get('mes', timezone.now().month))
    anio_sel = int(request.GET.get('anio', timezone.now().year))
    periodo_activo = Periodo.objects.filter(activo=True).first()
    socios_mes = Socio.objects.filter(estado='activo', fecha_nacimiento__month=mes_sel).order_by('fecha_nacimiento__day').prefetch_related('libretas')

    cumpleanos_data = []
    for socio in socios_mes:
        libretas_activas = socio.libretas.filter(estado='activa')
        if periodo_activo:
            libretas_activas = socio.libretas.filter(estado='activa', periodo=periodo_activo)
        n_libretas = libretas_activas.count()
        monto_regalo = n_libretas * 12
        egreso_registrado = Movimiento.objects.filter(
            categoria='cumpleanos',
            socio_ref__contains=socio.cedula,
            fecha__year=anio_sel,
            fecha__month=mes_sel,
        ).exists()
        cumpleanos_data.append({
            'socio': socio,
            'dia': socio.fecha_nacimiento.day,
            'libretas': n_libretas,
            'monto': monto_regalo,
            'egreso_registrado': egreso_registrado,
        })

    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    return render(request, 'core/cumpleanos.html', {
        'cumpleanos_data': cumpleanos_data,
        'mes_sel': mes_sel,
        'anio_sel': anio_sel,
        'mes_nombre': meses[mes_sel],
        'meses': list(enumerate(meses[1:], 1)),
        'total_mes': sum(d['monto'] for d in cumpleanos_data if not d['egreso_registrado']),
    })


@login_required
@require_roles('admin', 'tesorero', 'gerente')
def cumpleanos_registrar_egreso(request, socio_pk):
    socio = get_object_or_404(Socio, pk=socio_pk)
    periodo_activo = Periodo.objects.filter(activo=True).first()
    if request.method == 'POST':
        libretas_activas = socio.libretas.filter(estado='activa')
        if periodo_activo:
            libretas_activas = socio.libretas.filter(estado='activa', periodo=periodo_activo)
        n_libretas = libretas_activas.count()
        monto = Decimal(n_libretas * 12).quantize(Decimal('0.01'))
        mes = int(request.POST.get('mes', timezone.localdate().month))
        anio = int(request.POST.get('anio', timezone.localdate().year))
        fecha_movimiento = ddate(anio, mes, socio.fecha_nacimiento.day)
        otro_banco = request.POST.get('otro_banco') == '1'
        costo_operativo = Decimal('0.41') if otro_banco and monto > 0 else Decimal('0.00')
        monto_transferido = (monto - costo_operativo).quantize(Decimal('0.01'))
        movimiento = Movimiento.objects.create(
            tipo='egreso',
            categoria='cumpleanos',
            descripcion=f'Regalo cumpleaños - {socio.nombre_completo} - {n_libretas} libreta(s) x $12',
            monto=monto_transferido,
            fecha=fecha_movimiento,
            origen='manual',
            socio_ref=f'{socio.nombre_completo} ({socio.cedula})',
            registrado_por=request.user,
            observaciones='Transferencia a otro banco. Se descontó costo operativo de Transferencia a otro banco. Se descont? costo operativo de $0.41..41.' if otro_banco else '',
        )
        registrar_auditoria(request.user, 'tesoreria', 'egreso_cumpleanos', f'Se registró egreso por cumpleaños para {socio.nombre_completo}.', 'Movimiento', movimiento.pk)
        if otro_banco:
            Movimiento.objects.create(
                tipo='egreso',
                categoria='gasto_operativo',
                descripcion=f'Costo operativo transferencia cumpleaños - {socio.nombre_completo}',
                monto=costo_operativo,
                fecha=fecha_movimiento,
                origen='manual',
                socio_ref=f'{socio.nombre_completo} ({socio.cedula})',
                registrado_por=request.user,
                observaciones=f'Cobro bancario por transferencia de cumpleaños. Regalo original: ${monto}.',
            )
        messages.success(request, f'Cumpleaños registrado: ${monto_transferido} para el socio y ${costo_operativo} como costo operativo.') if otro_banco else messages.success(request, f'Egreso de ${monto} registrado.')
        notify_cumpleanos_transferido(socio, monto_transferido, costo_operativo)
    return redirect('cumpleanos_calendario')


@login_required
def busqueda_global(request):
    q = request.GET.get('q', '').strip()
    socios = []
    libretas = []
    creditos = []
    movimientos = []
    multas = []

    if q:
        socios = Socio.objects.filter(Q(nombres__icontains=q) | Q(apellidos__icontains=q) | Q(cedula__icontains=q))[:8]
        libretas = Libreta.objects.select_related('socio', 'periodo').filter(
            Q(numero__icontains=q) | Q(socio__nombres__icontains=q) | Q(socio__apellidos__icontains=q)
        )[:8]
        creditos = Credito.objects.select_related('socio', 'libreta').filter(
            Q(numero__icontains=q) | Q(socio__nombres__icontains=q) |
            Q(socio__apellidos__icontains=q) | Q(socio__cedula__icontains=q)
        )[:8]
        movimientos = Movimiento.objects.select_related('registrado_por').filter(
            Q(descripcion__icontains=q) | Q(comprobante__icontains=q) |
            Q(socio_ref__icontains=q) | Q(libreta_ref__icontains=q) | Q(credito_ref__icontains=q)
        )[:10]
        multas = Multa.objects.select_related('socio', 'libreta').filter(
            Q(socio__nombres__icontains=q) | Q(socio__apellidos__icontains=q) |
            Q(socio__cedula__icontains=q) | Q(descripcion__icontains=q)
        )[:8]

    return render(request, 'core/busqueda_global.html', {
        'q': q,
        'socios': socios,
        'libretas': libretas,
        'creditos': creditos,
        'movimientos': movimientos,
        'multas': multas,
        'estado_badge_credito': _badge_estado_credito,
    })


@login_required
def servir_comprobante(request, path):
    file_path = os.path.join(settings.MEDIA_ROOT, path)
    if not os.path.exists(file_path):
        return HttpResponse('Archivo no encontrado', status=404)
    content_type = 'application/octet-stream'
    if path.lower().endswith(('.jpg', '.jpeg')):
        content_type = 'image/jpeg'
    elif path.lower().endswith('.png'):
        content_type = 'image/png'
    elif path.lower().endswith('.webp'):
        content_type = 'image/webp'
    elif path.lower().endswith('.pdf'):
        content_type = 'application/pdf'
    return FileResponse(open(file_path, 'rb'), content_type=content_type)
